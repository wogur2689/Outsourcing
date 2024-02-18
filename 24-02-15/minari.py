import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By

import time

# 코드들
codes = ['복어', '아귀찜', '해물찜', '홍어', '등촌칼국수', '오리고기']
# 엑셀 url
fpath = r'C:\Users\PC\Desktop\크롤링\미나리_data.xlsx'
#엑셀 열기
wb = openpyxl.load_workbook(fpath)
#엑셀 워크시트 선택
ws = wb['등촌칼국수']
excel_row = 1

# service객체 생성
service = Service(executable_path='./chromedriver.exe')

# USB: usb_device_handle_win.cc:1046 Failed to read descriptor from node connection: 시스템에 부착된 장치가 작동하지 않습니다. (0x1F) 에러 방지
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

# 드라이버에 서비스 및 옵션 추가
driver = webdriver.Chrome(service = service, options=options)

#네이버 지도 이동
URL = "https://map.naver.com/p/search/등촌칼국수?searchType=place&c=15.00,0,0,0,dh"
driver.get(URL)

# selenium은 페이지를 직접 이동하면서 조작함. 지연시간 3초 주는게 좋음.
time.sleep(3)

#크롬창을 터미널에서 사용자 입력으로 닫는 코드
while True:
    
    for i in range(1, 51):
        #프레임으로 구성된 웹페이지라 탐색위치를 프레임 내부로 전환 CHC5F
        driver.switch_to.frame("searchIframe")
        time.sleep(3)

        # selenium의 by에서 검색방법(XPATH) 사용하여 해당 태그를 찾음.
        writing = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]')

        ws['A' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]/div/div/span[@class="YwYLL"]').text #식당명
        ws['B' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]/div/div/span[@class="YzBgS"]').text #분류
        ws['C' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/div/div/span[2]/a/span[1]').text #주소
    
        writing.click()
        time.sleep(2)

        # 이전 위치로 복귀
        driver.switch_to.default_content()
        time.sleep(1)

        #사이드바 프레임으로 이동(전화번호)
        driver.switch_to.frame("entryIframe")
        time.sleep(3)

        try:
            data = driver.find_element(By.CSS_SELECTOR, ".xlx7Q").text
        except NoSuchElementException:
            data = ''
            pass

        # 데이터 추가
        ws['D' + str(excel_row + 1)] = data
        ws['E' + str(excel_row + 1)] = '등촌칼국수'

        #저장
        wb.save(fpath)
        excel_row = int(excel_row + 1)

        # 이전 위치로 복귀
        driver.switch_to.default_content()
        time.sleep(2)

    # 다음페이지로 넘기기
    try:
        #프레임으로 구성된 웹페이지라 탐색위치를 프레임 내부로 전환
        driver.switch_to.frame("searchIframe")
        time.sleep(3)
        #페이지 이동
        page = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div[2]/a[7]")
        page.click()
        time.sleep(3)
        # 이전 위치로 복귀
        driver.switch_to.default_content()
        time.sleep(2)
    except NoSuchElementException: #단일페이지.
        pass

    
    # # selenium의 by에서 검색방법(CSS Selector) 사용하여 해당 태그를 찾음.
    # form = driver.find_element(By.CSS_SELECTOR, "input.input_search")
    # #해당 태그에 값 입력
    # form.send_keys('미나리')

    # # selenium의 by에서 검색방법(CSS Selector) 사용하여 해당 태그를 찾음.
    # btn = driver.find_element(By.CSS_SELECTOR, "input.input_search")
    # # 해당 태그를 클릭
    # btn.click()
    # # 검색 시간 소요 대비 지연시간
    # time.sleep(3)

    value = int(input("종료하시려면 0을 입력하세요 >>> "))
    if value == 0:
        print("2초후에 닫힙니다.")
        break

# 실행했던 크롬창 종료.
time.sleep(2)
driver.close()


# 미나리를 쓰는 식당 코드
# codes = [
#     '미나리',
#     '복어',
#     '아귀찜',
#     '해물찜',
#     '홍어',
#     '등촌칼국수',
#     '오리'
# ]

# for code in codes:
#     # naver 서버에 대화를 시도
#     response = requests.get(f'https://map.naver.com/p/search/{code}?c=18.81,0,0,0,dh') #fstring
#     # naver 에서 HTML을 줌.
#     html = response.text
#     # html 번역 선생님으로 수프 만듦
#     soup = BeautifulSoup(html, 'html.parser')
#     # id 값이 _nowVal인 놈 한개를 찾아냄(태그)
#     price = soup.select_one('#_nowVal').text # 결과
#     price = price.replace(',', '') #쉼표 제거
#     print(price)
#     ws[f'B{row}'] = int(price)
#     row = row + 1

# wb.save(fpath)