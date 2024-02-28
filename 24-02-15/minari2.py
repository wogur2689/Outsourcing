import openpyxl
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

import time

# 키워드
codes = ['복어', '아귀찜', '해물찜', '홍어', '등촌칼국수', '오리고기']

# 지역
busan = ['부산 강서구', '부산 사하구', '부산 중구', '부산 서구', '부산 동구', '부산 남구'
,'부산 부산진구', '부산 사상구', '부산 수영구', '부산 영도구', '부산 해운대구', '부산 북구', '부산 금정구', '부산 기장군']
Ulsan = ['울산 남구', '울산 동구', '울산 울주군', '울산 중구', '울산 북구']
Daegu = ['대구 달성군', '대구 달서구', '대구 남구', '대구 수성구', '대구 서구', '대구 중구', '대구 북구', '대구 동구', '대구 군위군']
GyeongsangnamDo = ['경상남도 밀양시', '경상남도 김해시', '경상남도 창원시', '경상남도 창녕군', '경상남도 합천군', '경상남도 거창군', '경상남도 함양군', '경상남도 산청군', '경상남도 의령군',
                    '경상남도 함안군', '경상남도 하동군', '경상남도 사천시', '경상남도 고성군', '경상남도 남해군',
                   '경상남도 거제시', '경상남도 통영시']
GyeongsangbukDo = ['경상북도 울진군', '경상북도 봉화군', '경상북도 영주시', '경상북도 영양군', '경상북도 안동시', '경상북도 영덕군', '경상북도 청송군',
                   '경상북도 경주시', '경상북도 청도군', '경상북도 경산시', '경상북도 영천시', '경상북도 의성군', '경상북도 안동시', '경상북도 예천군',
                   '경상북도 문경시', '경상북도 상주시', '경상북도 김천시', '경상북도 구미시', '경상북도 칠곡군', '경상북도 성주군', '경상북도 고령군']
Gwangju = ['광주 광산구', '광주 서구', '광주 북구', '광주 동구', '광주 남구']
JeollanamDo = ['전라남도 곡성군', '전라남도 구례군', '전라남도 광양시', '전라남도 여수시', '전라남도 순천시', '전라남도 고흥군', '전라남도 회순군',
               '전라남도 보성군', '전라남도 장흥군', '전라남도 영암군', '전라남도 강진군', '전라남도 해남군', '전라남도 진도군', '전라남도 영암군',
               '전라남도 무안군', '전라남도 나주시', '전라남도 함평군', '전라남도 신안군', '전라남도 영광군', '전라남도 장성군', '전라남도 담양군']
JeollabukDo = ['전라북도 고창군', '전라북도 정읍시', '전라북도 순창군', '전라북도 남원시', '전라북도 임실군', '전라북도 장수군', '전라북도 진안군',
               '전라북도 무주군', '전라북도 전주시', '전라북도 완주군', '전라북도 익산시', '전라북도 군산시', '전라북도 김제시', '전라북도 부안군']
jejuDo = ['제주도 제주시', '제주도 서귀포시']
chungbuk = ['충청북도 단양군', '충청북도 충주시', '충청북도 음성군', '충청북도 괴산군', '충청북도 증평군', '충청북도 진천군', '충청북도 청주시'
               '충청북도 옥천군', '충청북도 영동군']
chungnam = ['충청남도 태안군', '충청남도 서산시', '충청남도 당진시', '충청남도 아산시', '충청남도 천안시', '충청남도 예산군', '충청남도 홍성군',
            '충청남도 청양군', '충청남도 공주시', '충청남도 보령시', '충청남도 부여군', '충청남도 논산시', '충청남도 서천군', '충청남도 금산군']
gangwon = [ '강원도 고성군', '강원도 속초시', '강원도 양양군', '강원도 강릉시', '강원도 동해시' '강원도 삼척시', '강원도 태백시', '강원도 정선군'
           '강원도 영월군', '강원도 평창군', '강원도 원주시', '강원도 횡성군', '강원도 홍천군', '강원도 춘천시', '강원도 인제군', '강원도 양구군',
           '강원도 철원군'] 
daejen = ['대전 대덕구', '대전 동구', '대전 중구', '대전 서구', '대전 유성구'] 
# 광주, 대전, 대구, 인천, 세종, 부산, 울산, 제주도, 경상도, 강원도, 전라도는 시/구/군 분리 없음.
# 서울, 경기만 시/군/구 분리.
seoul = [ '서울 강서구', '서울 양천구', '서울 금천구', '서울 관악구', '서울 동작구', '서울 영등포구', '서울 마포구', '서울 서대문구',
        '서울 은평구', '서울 종로구', '서울 성북구', '서울 용산구', '서울 중구', '서울 강북구', '서울 도봉구', '서울 노원구', '서울 중랑구', '서울 동대문구',
        '서울 성동구' , '서울 광진구', '서울 서초구', '서울 강남구', '서울 송파구']
# '경기도 수원시', '경기도 성남시', '경기도 의정부시', '경기도 안양시', '경기도 부천시', '경기도 광명시', '경기도 평택시', '경기도 동두천시', '경기도 안산시',
#              '경기도 고양시', '경기도 과천시', '경기도 구리시', '경기도 남양주시',
Gyeonggido = [ '경기도 오산시', '경기도 시흥시', '경기도 군포시', '경기도 의왕시', '경기도 하남시', '경기도 용인시',
              '경기도 파주시', '경기도 이천시', '경기도 안성시', '경기도 김포시', '경기도 화성시', '경기도 광주시', '경기도 양주시', '경기도 포천시', '경기도 여주시', '경기도 연천군',
              '경기도 가평군', '경기도 양평군']
mapt = ['부산', '울산', '제주도', '세종', '인천', '대구', '대전', '광주', '전라북도', '전라남도', '경상북도', '경상남도', '강원도' '서울 강서구', '서울 양천구', '서울 금천구', '서울 관악구', '서울 동작구', '서울 영등포구', '서울 마포구', '서울 서대문구',
        '서울 은평구', '서울 종로구', '서울 성북구', '서울 용산구', '서울 중구', '서울 강북구', '서울 도봉구', '서울 노원구', '서울 중랑구', '서울 동대문구',
        '서울 성동구' , '서울 광진구', '서울 서초구', '서울 강남구', '서울 송파구']





# 엑셀 url
fpath = r'C:\Users\PC\Desktop\크롤링\미나리2_data.xlsx'
#엑셀 열기
wb = openpyxl.load_workbook(fpath)
#엑셀 워크시트 선택
ws = wb['복어']
excel_row = 2762

# service객체 생성
service = Service(executable_path='./chromedriver.exe')

# USB: usb_device_handle_win.cc:1046 Failed to read descriptor from node connection: 시스템에 부착된 장치가 작동하지 않습니다. (0x1F) 에러 방지
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])

# 드라이버에 서비스 및 옵션 추가
driver = webdriver.Chrome(service = service, options=options)

#네이버 지도 이동
URL = "https://map.naver.com/p/search/"
driver.get(URL)

# selenium은 페이지를 직접 이동하면서 조작함. 지연시간 3초 주는게 좋음.
time.sleep(3)

index = 0
value = 2
#크롬창을 터미널에서 사용자 입력으로 닫는 코드
while True:

    if value != 1:
        # selenium의 by에서 검색방법(CSS Selector) 사용하여 해당 태그를 찾음.
        form = driver.find_element(By.CSS_SELECTOR, "input.input_search")
        form.clear()
        for i in range(15):
            form.send_keys(Keys.BACKSPACE)

        time.sleep(2)
        #해당 태그에 값 입력
        keyword = str(Gyeonggido[index] + ' ' + codes[0])
        form.send_keys(keyword)
        time.sleep(1)
        # 검색
        form.send_keys(Keys.RETURN)
        # 검색 시간 소요 대비 지연시간
        time.sleep(4)

        '''
            식당으로 인식하는 경우와 그렇지 않은 경우가 있어 분기를 두개로 나눔.
            식당으로 인식하는 경우
        '''
        # selenium의 by에서 검색방법(CSS Selector) 사용하여 해당 태그를 찾음.
        tagCheck = True
        try:
            a = driver.find_element(By.CSS_SELECTOR, "div.place-pcmap-filter-area-wrapper")
            tagCheck = True
        except NoSuchElementException:
            tagCheck = False
            pass


    if tagCheck:
        for i in range(1, 51):
            # 맨 처음은 기본위치
            driver.switch_to.default_content()
            #프레임으로 구성된 웹페이지라 탐색위치를 프레임 내부로 전환 CHC5F
            driver.switch_to.frame("searchIframe")

            # selenium의 by에서 검색방법(XPATH) 사용하여 해당 태그를 찾음.
            # 만약 없을경우 생략.
            try:
                writing = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]')
            except NoSuchElementException:
                # 이전 위치로 복귀
                driver.switch_to.default_content()
                time.sleep(1)
                break

            ws['A' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]/div/div/span[@class="YwYLL"]').text #식당명
            ws['B' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/a[1]/div/div/span[@class="YzBgS"]').text #분류
            ws['C' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="ouxiq"]/div/div/span[2]/a/span[1]').text #주소
        
            writing.click()
            time.sleep(2)

            # 이전 위치로 복귀
            driver.switch_to.default_content()
            time.sleep(1)

            #사이드바 프레임으로 이동(전화번호)
            driver.switch_to.frame("entryIframe")
            time.sleep(2)

            try:
                data = driver.find_element(By.CSS_SELECTOR, ".xlx7Q").text
            except NoSuchElementException:
                data = ''
                pass

            # 데이터 추가
            ws['D' + str(excel_row + 1)] = data
            ws['E' + str(excel_row + 1)] = keyword

            #저장
            wb.save(fpath)
            excel_row = int(excel_row + 1)

            # 이전 위치로 복귀
            driver.switch_to.default_content()
            time.sleep(1)

    else:
        for i in range(1, 51):
            # 맨 처음은 기본위치
            driver.switch_to.default_content()
            #프레임으로 구성된 웹페이지라 탐색위치를 프레임 내부로 전환
            driver.switch_to.frame("searchIframe")
            time.sleep(1)

            # # 일정 수치만큼 스크롤 내리기
            # scroll_distance = 200  # 원하는 스크롤 거리 설정 (픽셀 단위)

            # # 현재 스크롤 위치 가져오기
            # current_scroll_position = driver.execute_script("return window.pageYOffset;")

            # # 새로운 스크롤 위치 계산
            # new_scroll_position = current_scroll_position + scroll_distance

            # # JavaScript를 사용하여 스크롤 실행
            # driver.execute_script("window.scrollTo(0, {});".format(new_scroll_position))
            # time.sleep(1)

            # selenium의 by에서 검색방법(XPATH) 사용하여 해당 태그를 찾음.
            # 만약 없을경우 생략.
            try:
                writing = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[@class="CHC5F"]/a[1]')
            except NoSuchElementException:
                # 이전 위치로 복귀
                driver.switch_to.default_content()
                time.sleep(1)
                break

            ws['A' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[@class="CHC5F"]/a[1]/div/div/span[@class="TYaxT"]').text #식당명
            ws['B' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[@class="CHC5F"]/a[1]/div/div/span[@class="KCMnt"]').text #분류
            # ws['C' + str(excel_row + 1)] = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[2]/div[1]/ul/li[' + str(i) + ']/div[1]/div[@class="CHC5F"]/div/div/span[2]/a/span[1]').text #주소 Pb4bU
        
            writing.click()
            time.sleep(2)

            # 이전 위치로 복귀
            driver.switch_to.default_content()
            time.sleep(1)

            #사이드바 프레임으로 이동(전화번호)
            driver.switch_to.frame("entryIframe")
            time.sleep(2)

            try:
                address = driver.find_element(By.CSS_SELECTOR, ".LDgIH").text
                ws['C' + str(excel_row + 1)] = address
            except NoSuchElementException:
                ws['C' + str(excel_row + 1)] = ''
                pass

            try:
                phone = driver.find_element(By.CSS_SELECTOR, ".xlx7Q").text
                ws['D' + str(excel_row + 1)] = phone
            except NoSuchElementException:
                ws['D' + str(excel_row + 1)] = ''
                pass

            # 데이터 추가
            ws['E' + str(excel_row + 1)] = keyword

            #저장
            wb.save(fpath)
            excel_row = int(excel_row + 1)

            # 이전 위치로 복귀
            driver.switch_to.default_content()
            time.sleep(1)

    value = int(input("종료하시려면 0을, 다음 페이지는 1을, 다음 키워드는 아무거나 입력하세요 >>> "))
    if value == 0:
        print("2초후에 닫힙니다.")
        break
    elif value == 1:
        print("다음 검색 시작")
        # 다음페이지로 넘기기
        try:
            #프레임으로 구성된 웹페이지라 탐색위치를 프레임 내부로 전환
            driver.switch_to.frame("searchIframe")
            time.sleep(1)
            #페이지 이동
            page = driver.find_element(By.XPATH, "/html/body/div[3]/div/div[2]/div[2]/a[7]")
            page.click()
            time.sleep(1)
            # 이전 위치로 복귀
            driver.switch_to.default_content()
            time.sleep(1)
        except NoSuchElementException: #단일페이지.
            pass
    else:
        index += 1
        #검색어 초기화

# 실행했던 크롬창 종료.
time.sleep(2)
driver.close()